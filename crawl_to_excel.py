import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

URL = "https://blabiepia.tistory.com/entry/TOEIC-LC-%EB%B9%88%EC%B6%9C-%ED%91%9C%ED%98%84-%EC%96%B4%ED%9C%98"

def crawl():
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0 Safari/537.36"}
    res = requests.get(URL, headers=headers)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "html.parser")

    content = soup.find("div", class_="article-view")

    if not content:
        raise ValueError("본문을 찾을 수 없습니다.")

    text = content.get_text(separator="\n")
    return text

def parse(text):
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        # "영어;한국어" 또는 "영어 ; 한국어" 형식
        for sep in [";", " ; ", ": ", " - ", "\t"]:
            if sep in line:
                parts = line.split(sep, 1)
                eng = parts[0].strip()
                kor = parts[1].strip()
                if eng and kor and len(eng) > 1:
                    rows.append((eng, kor))
                break
    return rows

def make_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "TOEIC LC 빈출 표현"

    # 헤더 스타일
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="2F5496")
    center = Alignment(horizontal="center", vertical="center")

    # 헤더
    headers = ["#", "영어 표현", "한국어 뜻"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 데이터
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (eng, kor) in enumerate(rows, 1):
        fill = PatternFill("solid", start_color="EEF2FF") if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for col, val in enumerate([i, eng, kor], 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # 열 너비
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40

    # 행 높이
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 18

    # 필터
    ws.auto_filter.ref = f"A1:C{len(rows)+1}"

    # 틀 고정
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"저장 완료: {path} ({len(rows)}개 항목)")

if __name__ == "__main__":
    print("블로그 크롤링 중...")
    text = crawl()
    print("데이터 파싱 중...")
    rows = parse(text)
    print(f"총 {len(rows)}개 항목 발견")
    out = r"C:\Users\User\Desktop\토익\TOEIC_LC_빈출표현.xlsx"
    make_excel(rows, out)
