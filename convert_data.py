import json
import re
import openpyxl

BASE = r"C:\Users\User\Desktop\토익"

def load_norangi():
    wb = openpyxl.load_workbook(f"{BASE}/노랭이 전면개정판.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        day, word, meaning, *_ = row
        if word and meaning:
            items.append({"day": str(day).strip(), "word": str(word).strip(), "meaning": str(meaning).strip()})
    wb.close()
    days = sorted(set(i["day"] for i in items), key=lambda d: int(re.search(r'\d+', d).group()))
    return {"id": "norangi", "name": "노랭이 전면개정판", "type": "vocab", "days": days, "items": items}

def load_lc():
    wb = openpyxl.load_workbook(f"{BASE}/TOEIC_LC_빈출표현.xlsx", read_only=True, data_only=True)
    ws = wb["TOEIC LC 빈출 표현"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        _, eng, kor = row[:3]
        if eng and kor:
            items.append({"word": str(eng).strip(), "meaning": str(kor).strip()})
    wb.close()
    return {"id": "lc", "name": "LC 빈출 표현", "type": "vocab", "items": items}

def load_blue():
    wb = openpyxl.load_workbook(f"{BASE}/토익어휘 파란색책 2500문제.xlsx", read_only=True, data_only=True)
    ws = wb["파란색 교재 어휘"]
    items = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        word = row[0]
        meaning = row[1] if len(row) > 1 else None
        if word and meaning:
            items.append({"word": str(word).strip(), "meaning": str(meaning).strip()})
    wb.close()
    return {"id": "blue", "name": "파란색책 어휘 2500", "type": "vocab", "items": items}

def load_grammar():
    wb = openpyxl.load_workbook(f"{BASE}/토익_문법정리.xlsx", read_only=True, data_only=True)
    ws = wb["토익문법"]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat, point, desc, example, trap, freq = row[:6]
        if point and desc:
            items.append({
                "category": str(cat).strip() if cat else "",
                "point": str(point).strip(),
                "description": str(desc).strip(),
                "example": str(example).strip() if example else "",
                "trap": str(trap).strip() if trap else "",
                "frequency": str(freq).strip() if freq else ""
            })
    wb.close()
    return {"id": "grammar", "name": "토익 문법정리", "type": "grammar", "items": items}

if __name__ == "__main__":
    sources = [load_norangi(), load_lc(), load_blue(), load_grammar()]
    for s in sources:
        print(f"  {s['name']}: {len(s['items'])}개")

    data_js = "const TOEIC_DATA = " + json.dumps(sources, ensure_ascii=False, indent=None) + ";\n"

    with open(f"{BASE}/data.js", "w", encoding="utf-8") as f:
        f.write(data_js)
    print(f"data.js 생성 완료 ({len(data_js)//1024}KB)")
